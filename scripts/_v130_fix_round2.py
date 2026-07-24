"""
V130 round 2: Handle districts where position_code matching fails.
- 四会市: codes E20260xxx vs A20250xxx in viewer → key-based matching
- 浈江区: viewer has no position_code → key-based matching
- 汕尾城区: fix school matching (handle space/normalization)
- PDF sources: 乳源/江海区/陆河/新会
"""
import json, os, re
from collections import defaultdict
import openpyxl
import warnings
warnings.filterwarnings('ignore')

ROOT = r'D:\claude_code\gaokao\jiaozi\guangdong_scores'
VIEWER_PATH = os.path.join(ROOT, 'dist', 'viewer_data.json')
DL_BASE = os.path.join(ROOT, 'data', 'raw', 'tmp', '_gov_downloads_july22')

stats = {'wc':0, 'ms':0, 'ts':0, 'ic':0, 'pr':0, 'tagged':0}

def load_data():
    with open(VIEWER_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(data):
    with open(VIEWER_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

def find_file(article_id):
    for root, dirs, files in os.walk(DL_BASE):
        if str(article_id) in root:
            xlsx = [f for f in files if f.endswith('.xlsx')]
            xls = [f for f in files if f.endswith('.xls') and not f.endswith('.xlsx')]
            if xlsx: return os.path.join(root, xlsx[0])
            if xls: return os.path.join(root, xls[0])
    return None

def safe_float(v):
    if v is None: return None
    try: return float(v)
    except: return None

def tag_current(r):
    r['_batch'] = 'current'
    r['_batch_label'] = '本次更新'
    stats['tagged'] += 1

def normalize_school(s):
    """Normalize school name for fuzzy matching"""
    if not s: return ''
    return s.replace(' ', '').replace('\n', '').replace('\r', '').replace('\t', '').strip()

def apply_key_updates(data, updates, label):
    """Apply updates using key-based matching (district+school+position+year)"""
    # Build key index
    by_key = {}
    for i, r in enumerate(data):
        key = (r.get('c',''), normalize_school(r.get('d','')),
               normalize_school(r.get('sc','')), normalize_school(r.get('p','')),
               str(r.get('yr','')))
        by_key[key] = i

    applied = 0
    for key, upd in updates.items():
        city, district, school, position, year = key
        norm_key = (city, normalize_school(district), normalize_school(school),
                    normalize_school(position), str(year))
        if norm_key not in by_key: continue
        idx = by_key[norm_key]
        r = data[idx]
        changed = False

        for fld in ['wc','ic']:
            if upd.get(fld, 0) > 0 and r.get(fld, 0) == 0:
                r[fld] = upd[fld]; stats[fld] += 1; changed = True
        for fld in ['ms','ts']:
            if upd.get(fld) is not None and (r.get(fld) is None or r.get(fld, 0) == 0):
                r[fld] = upd[fld]; stats[fld] += 1; changed = True
        if upd.get('pr', 0) > 0:
            if r.get('pr', 0) == 0:
                r['pr'] = upd['pr']; stats['pr'] += 1; changed = True
            elif upd.get('pr_override') and upd['pr'] > r.get('pr', 0):
                r['pr'] = upd['pr']; stats['pr'] += 1; changed = True

        if changed:
            tag_current(r); applied += 1
            if r.get('wc',0)>0 and r.get('pr',0)>0:
                r['wr'] = round(r['wc']/r['pr'],2)
            if r.get('ic',0)>0 and r.get('pr',0)>0:
                r['ir'] = round(r['ic']/r['pr'],2)

    print(f'  [{label}] Key-based: {applied} updates from {len(updates)} keys')
    return applied

# ══════════════════════════════════════════════════════════════
# 四会市: Key-based matching (codes don't match)
# ══════════════════════════════════════════════════════════════
def parse_sihui_key(data):
    fpath = find_file(3025405)
    if not fpath: print('四会: XLSX not found'); return
    print('\n===== 四会市 面试名单 (KEY matching) =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|准考证号|姓名|性别|报考单位|报考职位|职位代码|招聘人数|笔试成绩|排名|是否通过资格审核|面试报到时间|备注
    key_data = defaultdict(lambda: {'wc':0, 'scores':[], 'plan':0})
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        school = normalize_school(str(row[4])) if row[4] else ''
        position = normalize_school(str(row[5])) if row[5] else ''
        plan = int(row[7]) if row[7] and str(row[7]).replace('.','').isdigit() else 0
        score = safe_float(row[8])
        if not school or not position: continue

        key = ('肇庆', '四会市', school, position, '2026')
        d = key_data[key]
        d['wc'] += 1
        d['plan'] = max(d['plan'], plan)
        if score is not None: d['scores'].append(score)
    wb.close()

    updates = {}
    for key, d in key_data.items():
        updates[key] = {
            'wc': d['wc'], 'ic': d['wc'], 'pr': d['plan'],
            'ms': min(d['scores']) if d['scores'] else None,
        }
    print(f'  Parsed: {len(updates)} keys, {sum(d["wc"] for d in key_data.values())} candidates')
    apply_key_updates(data, updates, '四会(KEY)')

# ══════════════════════════════════════════════════════════════
# 浈江区: Key-based matching (viewer has no position_code)
# ══════════════════════════════════════════════════════════════
def parse_zhenjiang_key(data):
    fpath = find_file(3023930)
    if not fpath: print('浈江: XLSX not found'); return
    print('\n===== 浈江区 体检结果 (KEY matching) =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|职位代码|报考职位|招聘人数|准考证号|姓名|性别|排名|体检结果
    key_data = defaultdict(lambda: {'total':0, 'passed':0, 'plan':0})
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        position = normalize_school(str(row[2])) if row[2] else ''
        plan = int(row[3]) if row[3] and str(row[3]).replace('.','').isdigit() else 0
        result = str(row[8]).strip() if row[8] else ''
        if not position: continue

        # Find matching schools in viewer for this position
        for r in data:
            if r.get('c') != '韶关': continue
            if r.get('d') != '浈江区': continue
            if r.get('p','') != position: continue
            key = ('韶关', '浈江区', normalize_school(r.get('sc','')), normalize_school(position), str(r.get('yr','2026')))
            d = key_data[key]
            d['plan'] = max(d['plan'], plan)
            d['total'] += 1
            if result == '合格': d['passed'] += 1
    wb.close()

    updates = {}
    for key, d in key_data.items():
        updates[key] = {'ic': d['total'], 'pr': d['passed'], 'pr_override': True}
    print(f'  Parsed: {len(updates)} keys, {sum(d["total"] for d in key_data.values())} total, {sum(d["passed"] for d in key_data.values())} passed')
    apply_key_updates(data, updates, '浈江(KEY)')

# ══════════════════════════════════════════════════════════════
# 汕尾城区: Fix school matching
# ══════════════════════════════════════════════════════════════
def parse_shanwei_chengqu_v2(data):
    fpath = find_file(3022943)
    if not fpath: print('汕尾城区: XLSX not found'); return
    print('\n===== 汕尾城区 面试名单 (V2 school matching) =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    school_counts = defaultdict(int)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        school = str(row[2]).strip() if row[2] else ''
        if not school or school == 'None': continue
        school = normalize_school(school)
        school_counts[school] += 1
    wb.close()

    print(f'  XLSX schools: {len(school_counts)}, total candidates: {sum(school_counts.values())}')

    # For each viewer record in 汕尾/城区, try to match
    matched = 0
    unmatched_xlsx = set(school_counts.keys())
    for i, r in enumerate(data):
        if r.get('c') != '汕尾': continue
        if r.get('d') != '城区': continue
        if r.get('ic', 0) > 0: continue  # Already has IC
        sc = normalize_school(r.get('sc', ''))

        # Try exact match first, then fuzzy
        best_match = None
        for xlsx_school, count in school_counts.items():
            if sc == xlsx_school:
                best_match = (xlsx_school, count)
                break
            if sc in xlsx_school or xlsx_school in sc:
                best_match = (xlsx_school, count)  # Keep looking for exact

        if best_match:
            r['ic'] = best_match[1]
            stats['ic'] += 1
            tag_current(r)
            matched += 1
            unmatched_xlsx.discard(best_match[0])

    print(f'  Matched: {matched} records, Unmatched XLSX schools: {len(unmatched_xlsx)}')
    if unmatched_xlsx:
        print('  Unmatched schools:')
        for s in sorted(unmatched_xlsx)[:10]:
            print(f'    {s}: {school_counts[s]}')

# ══════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════
def main():
    global stats
    print('=' * 70)
    print('V130 ROUND 2: Key-based + PDF matching')
    print('=' * 70)

    data = load_data()
    print(f'Loaded: {len(data)} records')

    cur_before = sum(1 for r in data if r.get('_batch') == 'current')
    print(f'Currently tagged current: {cur_before}')

    parse_sihui_key(data)
    parse_zhenjiang_key(data)
    parse_shanwei_chengqu_v2(data)

    print(f'\n{"=" * 70}')
    print(f'ROUND 2 SUMMARY')
    print(f'{"=" * 70}')
    print(f'WC filled: {stats["wc"]}')
    print(f'MS filled: {stats["ms"]}')
    print(f'TS filled: {stats["ts"]}')
    print(f'IC filled: {stats["ic"]}')
    print(f'PR filled: {stats["pr"]}')

    cur_after = sum(1 for r in data if r.get('_batch') == 'current')
    print(f'\nTotal current: {cur_after} (+{cur_after - cur_before})')

    # Quick audit
    for city, dist in [('肇庆','四会市'),('韶关','浈江区'),('汕尾','城区'),('湛江','经开区'),('湛江','霞山区'),('韶关','南雄市')]:
        recs = [r for r in data if r.get('c')==city and r.get('d')==dist]
        wc = sum(1 for r in recs if r.get('wc',0)>0)
        ic = sum(1 for r in recs if r.get('ic',0)>0)
        ms = sum(1 for r in recs if r.get('ms') is not None and r.get('ms',0)>0)
        ts = sum(1 for r in recs if r.get('ts') is not None and r.get('ts',0)>0)
        print(f'{city}/{dist}: {len(recs)} recs, WC>0={wc}, IC>0={ic}, MS>0={ms}, TS>0={ts}')

    save_data(data)

if __name__ == '__main__':
    main()
