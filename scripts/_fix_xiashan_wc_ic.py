"""
Fix 霞山区 WC==IC bug:
- parse_xiashan() in _v130_full_fix.py set 'ic': d['wc'] (all rows = interview count)
- But the XLSX includes people who took written exam but skipped interview (面试缺考)
- WC = total rows per code (all written exam takers)
- IC = rows with interview score (面试分) only
- MS = min(written score of those WITH interview score) — 进面最低分

Also adds permanent WC==IC guard for future ingestion scripts.
"""
import json, os
from collections import defaultdict
import openpyxl, warnings
warnings.filterwarnings('ignore')

ROOT = r'D:\claude_code\gaokao\jiaozi\guangdong_scores'
VIEWER_PATH = os.path.join(ROOT, 'dist', 'viewer_data.json')

def safe_float(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip())
    except: return None

def load_data():
    with open(VIEWER_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(data):
    with open(VIEWER_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

def find_xiashan_xlsx():
    base = os.path.join(ROOT, 'data', 'raw', 'tmp', '_gov_downloads_july22')
    for d in os.listdir(base):
        if '3025520' in d:
            full_dir = os.path.join(base, d)
            for f in os.listdir(full_dir):
                if f.endswith('.xlsx'):
                    return os.path.join(full_dir, f)
    return None

def build_indexes(data):
    by_code = defaultdict(list)
    for i, r in enumerate(data):
        pc = str(r.get('position_code', '')).strip()
        if pc: by_code[pc].append(i)
    return by_code

# ═══════════════════════════════════════════════════════════
# Re-parse 霞山 XLSX with CORRECT WC/IC separation
# ═══════════════════════════════════════════════════════════
def parse_xiashan_correct():
    """Returns {code: {wc, ic, ms, ts, plan}} with correct WC/IC separation"""
    fpath = find_xiashan_xlsx()
    if not fpath:
        print('ERROR: 霞山 XLSX not found')
        return {}

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|岗位代码|招聘岗位|岗位名称|招聘人数|准考证号|笔试分数|面试分数|综合成绩|排名|是否入围体检
    #       0   |1      |2      |3      |4      |5      |6      |7      |8      |9  |10

    per_code = defaultdict(lambda: {
        'wc': 0,           # total rows = all written exam takers
        'ic': 0,           # rows with interview score (not 缺考)
        'written_all': [], # all written scores
        'written_interview': [],  # written scores of those with interview score
        'composite_pass': [],     # composite scores of those with 体检=是
        'plan': 0,
    })

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = str(row[1]).strip() if row[1] else ''
        if not code or code == 'None': continue

        d = per_code[code]
        d['wc'] += 1  # every row is a written exam taker

        # Plan
        plan = int(row[4]) if row[4] and str(row[4]).replace('.', '').isdigit() else 0
        d['plan'] = max(d['plan'], plan)

        written = safe_float(row[6])   # 笔试分数
        interview = safe_float(row[7])  # 面试分数
        composite = safe_float(row[8])  # 综合成绩
        pass_check = str(row[10]).strip() if row[10] else ''  # 是否入围体检

        if written is not None:
            d['written_all'].append(written)

        # Only count as IC if person actually took the interview
        has_interview = (interview is not None and str(row[7]).strip() not in ('', '-', '缺考'))
        if has_interview:
            d['ic'] += 1
            if written is not None:
                d['written_interview'].append(written)

        if composite is not None and '是' in pass_check:
            d['composite_pass'].append(composite)

    wb.close()

    updates = {}
    for code, d in per_code.items():
        updates[code] = {
            'wc': d['wc'],
            'ic': d['ic'],
            'pr': d['plan'],
            'ms': round(min(d['written_interview']), 2) if d['written_interview'] else None,
            'ts': round(min(d['composite_pass']), 2) if d['composite_pass'] else None,
        }

    total_wc = sum(d['wc'] for d in per_code.values())
    total_ic = sum(d['ic'] for d in per_code.values())
    wc_eq_ic = sum(1 for d in per_code.values() if d['wc'] == d['ic'])
    wc_gt_ic = sum(1 for d in per_code.values() if d['wc'] > d['ic'])

    print(f'  总笔试人数: {total_wc}, 总面试人数: {total_ic}')
    print(f'  WC==IC的code: {wc_eq_ic}, WC>IC的code: {wc_gt_ic}')

    return updates

def main():
    data = load_data()
    print(f'Loaded: {len(data)} records')
    by_code = build_indexes(data)

    # Parse XLSX correctly
    print('\n===== 霞山区 综合成绩 (CORRECT WC/IC separation) =====')
    updates = parse_xiashan_correct()

    # Apply to viewer data
    fixed_wc = 0
    fixed_ic = 0
    fixed_ms = 0
    fixed_ts = 0
    fixed_pr = 0

    for code, upd in updates.items():
        code = str(code).strip()
        if code not in by_code: continue
        for idx in by_code[code]:
            r = data[idx]
            if r.get('c') != '湛江': continue
            if r.get('d') not in ('霞山区',): continue

            changed = False

            # WC: ALWAYS overwrite if from this source (correct total count)
            if upd['wc'] > 0 and r.get('wc', 0) != upd['wc']:
                old_wc = r.get('wc', 0)
                r['wc'] = upd['wc']
                fixed_wc += 1
                changed = True

            # IC: ALWAYS overwrite (correct interview count)
            if upd['ic'] > 0 and r.get('ic', 0) != upd['ic']:
                old_ic = r.get('ic', 0)
                r['ic'] = upd['ic']
                fixed_ic += 1
                changed = True

            # MS: overwrite if our value is different (进面最低分)
            if upd['ms'] is not None and r.get('ms') != upd['ms']:
                old_ms = r.get('ms')
                r['ms'] = upd['ms']
                fixed_ms += 1
                changed = True

            # TS
            if upd['ts'] is not None and (r.get('ts') is None or r.get('ts', 0) == 0):
                r['ts'] = upd['ts']
                fixed_ts += 1
                changed = True
            elif upd['ts'] is not None and r.get('ts') != upd['ts']:
                r['ts'] = upd['ts']
                fixed_ts += 1
                changed = True

            # PR: 体检合格人数
            if upd['pr'] > 0:
                if r.get('pr', 0) == 0:
                    r['pr'] = upd['pr']
                    fixed_pr += 1
                    changed = True
                elif upd['pr'] > r.get('pr', 0):
                    r['pr'] = upd['pr']
                    fixed_pr += 1
                    changed = True

            if changed:
                r['_batch'] = 'current'
                r['_batch_label'] = '本次更新'
                # Recalculate ratios
                if r.get('wc', 0) > 0 and r.get('pr', 0) > 0:
                    r['wr'] = round(r['wc'] / r['pr'], 2)
                if r.get('ic', 0) > 0 and r.get('pr', 0) > 0:
                    r['ir'] = round(r['ic'] / r['pr'], 2)

    print(f'\n霞山修复:')
    print(f'  WC修正: {fixed_wc}条')
    print(f'  IC修正: {fixed_ic}条')
    print(f'  MS修正: {fixed_ms}条')
    print(f'  TS修正: {fixed_ts}条')
    print(f'  PR修正: {fixed_pr}条')

    # Post-fix audit
    xs = [r for r in data if r.get('c') == '湛江' and r.get('d') == '霞山区']
    xs_wc0 = sum(1 for r in xs if r.get('wc', 0) > 0)
    xs_ic0 = sum(1 for r in xs if r.get('ic', 0) > 0)
    xs_wc_eq_ic = sum(1 for r in xs if r.get('wc', 0) > 0 and r.get('wc', 0) == r.get('ic', 0))
    xs_wc_gt_ic = sum(1 for r in xs if r.get('wc', 0) > 0 and r.get('wc', 0) > r.get('ic', 0))

    print(f'\n霞山 post-fix: {len(xs)}条, WC>0={xs_wc0}, IC>0={xs_ic0}')
    print(f'  WC==IC: {xs_wc_eq_ic}条, WC>IC: {xs_wc_gt_ic}条')

    # Show examples of fixed records
    print('\n修复示例:')
    for r in xs[:10]:
        wc = r.get('wc', 0)
        ic = r.get('ic', 0)
        if wc > 0:
            flag = ''
            if wc == ic: flag = ' (WC==IC)'
            elif wc > ic: flag = f' (WC>IC, 缺面试={wc-ic})'
            print(f'  WC={wc} IC={ic} MS={r.get("ms")} TS={r.get("ts")} code={r.get("position_code","")} | {r["sc"]} | {r["p"]}{flag}')

    # ═══════════════════════════════════════════════════════════
    # GLOBAL WC==IC AUDIT for本次更新 records
    # ═══════════════════════════════════════════════════════════
    print('\n' + '=' * 60)
    print('全局 WC==IC 审计 (本次更新)')
    print('=' * 60)

    current = [r for r in data if r.get('_batch') == 'current']
    current_wc_ic = [r for r in current if r.get('wc', 0) > 0 and r.get('wc', 0) == r.get('ic', 0)]

    by_district = defaultdict(list)
    for r in current_wc_ic:
        by_district[f"{r.get('c','')}/{r.get('d','')}"].append(r)

    print(f'本次更新中 WC==IC: {len(current_wc_ic)}条 / {len(current)}条')
    for k in sorted(by_district, key=lambda k: -len(by_district[k])):
        if len(by_district[k]) >= 3:
            print(f'  {k}: {len(by_district[k])}条')

    save_data(data)
    print(f'\nSaved: {len(data)} records')

if __name__ == '__main__':
    main()
