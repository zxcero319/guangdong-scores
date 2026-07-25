"""
V130 COMPREHENSIVE FIX:
1. Re-parse ALL v130 XLSX/XLS sources to fill WC/IC/MS/TS/PR where missing
2. Fix 经开区 MS: use interview list scores (进面最低分), NOT all-test-taker min
3. Fill 经开区 IC from interview list (2205044.xlsx)
4. Fill IC for 汕尾市直/城区/普宁
5. Fill WC/MS/TS/IC for 四会/浈江/新丰/连州
6. All filled records tagged _batch='current'
"""
import json, os
from collections import defaultdict
import openpyxl
import warnings
warnings.filterwarnings('ignore')

ROOT = r'D:\claude_code\gaokao\jiaozi\guangdong_scores'
VIEWER_PATH = os.path.join(ROOT, 'dist', 'viewer_data.json')
DL_BASE = os.path.join(ROOT, 'data', 'raw', 'tmp', '_gov_downloads_july22')

stats = {'wc':0, 'ms':0, 'ms_corrected':0, 'ts':0, 'ic':0, 'pr':0, 'tagged':0}

def load_data():
    with open(VIEWER_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(data):
    with open(VIEWER_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    print(f'  Saved: {len(data)} records')

def find_file(article_id):
    for root, dirs, files in os.walk(DL_BASE):
        if str(article_id) in root:
            xlsx = [f for f in files if f.endswith('.xlsx')]
            xls = [f for f in files if f.endswith('.xls') and not f.endswith('.xlsx')]
            if xlsx: return os.path.join(root, xlsx[0])
            if xls: return os.path.join(root, xls[0])
    return None

def safe_float(v):
    if v is None: return None
    try: return float(v)
    except: return None

def build_indexes(data):
    by_code = defaultdict(list)
    by_key = {}
    for i, r in enumerate(data):
        pc = str(r.get('position_code', '')).strip()
        if pc: by_code[pc].append(i)
        key = (r.get('c',''), r.get('d',''), r.get('sc',''), r.get('p',''), str(r.get('yr','')))
        by_key[key] = i
    return by_code, by_key

def tag_current(r):
    r['_batch'] = 'current'
    r['_batch_label'] = '本次更新'
    stats['tagged'] += 1

def apply_code_updates(data, by_code, updates, label):
    """Apply WC/MS/TS/IC/PR updates to viewer records by position_code"""
    applied = 0
    wc_eq_ic_warns = 0
    for code, upd in updates.items():
        code = str(code).strip()
        if code not in by_code: continue
        for idx in by_code[code]:
            r = data[idx]
            changed = False

            # ═══ WC==IC PERMANENT GUARD ═══
            # Check if applying this update would result in WC==IC
            new_wc = upd.get('wc', 0) if upd.get('wc', 0) > 0 else r.get('wc', 0)
            new_ic = upd.get('ic', 0) if upd.get('ic', 0) > 0 else r.get('ic', 0)
            if new_wc > 0 and new_wc == new_ic:
                # Only warn if both WC and IC are being set from this update
                if upd.get('wc', 0) > 0 and upd.get('ic', 0) > 0:
                    if wc_eq_ic_warns < 3:
                        print(f'  [{label}] WC==IC GUARD: code={code} wc=ic={new_wc} '
                              f'| 确认来源为完整笔试名单(含面试缺考行)? 综合成绩表可WC==IC.')
                    wc_eq_ic_warns += 1

            if upd.get('wc', 0) > 0 and r.get('wc', 0) == 0:
                r['wc'] = upd['wc']; stats['wc'] += 1; changed = True
            if upd.get('ms') is not None and (r.get('ms') is None or r.get('ms', 0) == 0):
                r['ms'] = upd['ms']; stats['ms'] += 1; changed = True
            if upd.get('ts') is not None and (r.get('ts') is None or r.get('ts', 0) == 0):
                r['ts'] = upd['ts']; stats['ts'] += 1; changed = True
            if upd.get('ic', 0) > 0 and r.get('ic', 0) == 0:
                r['ic'] = upd['ic']; stats['ic'] += 1; changed = True
            if upd.get('pr', 0) > 0:
                if r.get('pr', 0) == 0:
                    r['pr'] = upd['pr']; stats['pr'] += 1; changed = True
                elif upd.get('pr_override') and upd['pr'] > r.get('pr', 0):
                    # 体检合格 > 计划数
                    r['pr'] = upd['pr']; stats['pr'] += 1; changed = True

            if changed:
                tag_current(r)
                applied += 1

            # Recalculate ratios
            if r.get('wc', 0) > 0 and r.get('pr', 0) > 0:
                r['wr'] = round(r['wc'] / r['pr'], 2)
            if r.get('ic', 0) > 0 and r.get('pr', 0) > 0:
                r['ir'] = round(r['ic'] / r['pr'], 2)

    if wc_eq_ic_warns > 0:
        print(f'  [{label}] WC==IC guard: {wc_eq_ic_warns} codes where WC==IC')
    print(f'  [{label}] {applied} record updates from {len(updates)} codes')
    return applied

def correct_ms(data, by_code, correct_ms_map, label):
    """CORRECT MS values: overwrite existing MS with correct values"""
    corrected = 0
    for code, correct_val in correct_ms_map.items():
        code = str(code).strip()
        if code not in by_code: continue
        for idx in by_code[code]:
            r = data[idx]
            old_ms = r.get('ms')
            if old_ms != correct_val:
                r['ms'] = correct_val
                stats['ms_corrected'] += 1
                corrected += 1
                tag_current(r)
    print(f'  [{label}] MS corrected: {corrected} records')

# ══════════════════════════════════════════════════════════════
# Source 1: 霞山区 综合成绩 (2204552.xlsx) — WC+MS+TS+IC+PR
# ══════════════════════════════════════════════════════════════
def parse_xiashan(data, by_code):
    fpath = find_file(3025520)
    if not fpath: print('霞山区: XLSX not found'); return
    print('\n===== 霞山区 综合成绩 =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|岗位代码|招聘岗位|岗位名称|招聘人数|准考证号|笔试分数|面试分数|综合成绩|排名|是否入围体检
    pos_data = defaultdict(lambda: {'wc':0, 'written':[], 'composite':[], 'pass_composite':[], 'school':'', 'position':'', 'plan':0})
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = str(row[1]).strip() if row[1] else ''
        if not code or code == 'None': continue
        d = pos_data[code]
        d['school'] = str(row[2]).strip() if row[2] else ''
        d['position'] = str(row[3]).strip() if row[3] else ''
        d['plan'] = max(d['plan'], int(row[4]) if row[4] and str(row[4]).replace('.','').isdigit() else 0)
        d['wc'] += 1
        written = safe_float(row[6])
        composite = safe_float(row[8])
        pass_check = str(row[10]).strip() if row[10] else ''
        if written is not None: d['written'].append(written)
        if composite is not None:
            d['composite'].append(composite)
            if '是' in pass_check: d['pass_composite'].append(composite)
    wb.close()

    updates = {}
    for code, d in pos_data.items():
        updates[code] = {
            'wc': d['wc'], 'ic': d['wc'],
            'pr': d['plan'],
            'ms': min(d['written']) if d['written'] else None,
            'ts': min(d['pass_composite']) if d['pass_composite'] else None,
        }
    print(f'  Parsed: {len(updates)} positions, {sum(d["wc"] for d in pos_data.values())} candidates')
    apply_code_updates(data, by_code, updates, '霞山')

# ══════════════════════════════════════════════════════════════
# Source 2: 南雄市 总成绩 (2863744.xlsx) — WC+MS+TS+IC+PR
# ══════════════════════════════════════════════════════════════
def parse_nanxiong(data, by_code):
    fpath = find_file(3018354)
    if not fpath: print('南雄: XLSX not found'); return
    print('\n===== 南雄市 总成绩 =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb['Sheet1']
    # Cols: 序号|准考证号|岗位代码|岗位名称|招聘单位|招聘人数|笔试成绩|面试成绩|总成绩|排名|是否进入体检|备注
    pos_data = defaultdict(lambda: {'wc':0, 'written':[], 'total':[], 'pass_total':[], 'school':'', 'position':'', 'plan':0})
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        code = str(row[2]).strip() if row[2] else ''
        if not code or code == 'None': continue
        d = pos_data[code]
        d['position'] = str(row[3]).strip() if row[3] else ''
        d['school'] = str(row[4]).strip() if row[4] else ''
        d['plan'] = max(d['plan'], int(row[5]) if row[5] and str(row[5]).replace('.','').isdigit() else 0)
        d['wc'] += 1
        written = safe_float(row[6])
        total = safe_float(row[8])
        pass_check = str(row[10]).strip() if row[10] else ''
        if written is not None: d['written'].append(written)
        if total is not None:
            d['total'].append(total)
            if '是' in pass_check: d['pass_total'].append(total)
    wb.close()

    updates = {}
    for code, d in pos_data.items():
        updates[code] = {
            'wc': d['wc'], 'ic': d['wc'],
            'pr': d['plan'],
            'ms': min(d['written']) if d['written'] else None,
            'ts': min(d['pass_total']) if d['pass_total'] else None,
        }
    print(f'  Parsed: {len(updates)} positions, {sum(d["wc"] for d in pos_data.values())} candidates')
    apply_code_updates(data, by_code, updates, '南雄')

# ══════════════════════════════════════════════════════════════
# Source 3: 经开区 笔试成绩 (2203725.xlsx) — WC only
# ══════════════════════════════════════════════════════════════
def parse_jkq_written(data, by_code):
    """Fill WC from written exam list. MS is NOT filled here — use interview list."""
    fpath = find_file(3021062)
    if not fpath: print('经开区笔试: XLSX not found'); return
    print('\n===== 经开区 笔试成绩 (WC only) =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|岗位代码|姓名|准考证号|笔试成绩|岗位排名
    pos_wc = defaultdict(int)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = str(row[1]).strip() if row[1] else ''
        if not code or code == 'None': continue
        pos_wc[code] += 1
    wb.close()

    updates = {code: {'wc': wc} for code, wc in pos_wc.items()}
    print(f'  Parsed: {len(updates)} codes, {sum(pos_wc.values())} candidates')
    apply_code_updates(data, by_code, updates, '经开区笔试→WC')

# ══════════════════════════════════════════════════════════════
# Source 4: 经开区 面试名单 (2205044.xlsx) — IC + CORRECT MS
# ══════════════════════════════════════════════════════════════
def parse_jkq_interview(data, by_code):
    """Fill IC from interview list. Also CORRECT MS = min(面试名单 grades).
    MS here is the 进面最低分 (interview cutoff), not all-test-taker min."""
    fpath = find_file(3028691)
    if not fpath: print('经开区面试: XLSX not found'); return
    print('\n===== 经开区 面试名单 (IC + correct MS) =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Header: 序号|岗位代码|准考证号|姓名|成绩|排名|备注
    # "成绩" = 笔试成绩 of interview entrants
    pos_data = defaultdict(lambda: {'ic': 0, 'scores': []})
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = str(row[1]).strip() if row[1] else ''
        if not code or code == 'None': continue
        d = pos_data[code]
        d['ic'] += 1
        score = safe_float(row[4])
        if score is not None: d['scores'].append(score)
    wb.close()

    print(f'  Parsed: {len(pos_data)} codes, {sum(d["ic"] for d in pos_data.values())} candidates')

    # 1. Fill IC
    ic_updates = {code: {'ic': d['ic']} for code, d in pos_data.items()}
    apply_code_updates(data, by_code, ic_updates, '经开区面试→IC')

    # 2. CORRECT MS: overwrite with min from interview list
    ms_correct = {}
    for code, d in pos_data.items():
        if d['scores']:
            ms_correct[code] = round(min(d['scores']), 2)
    correct_ms(data, by_code, ms_correct, '经开区面试→MS(进面分)')

    # Show comparison
    for code in sorted(pos_data.keys())[:5]:
        d = pos_data[code]
        new_ms = round(min(d['scores']), 2) if d['scores'] else None
        # Find current MS in data
        old_ms = None
        if code in by_code:
            for idx in by_code[code]:
                old_ms = data[idx].get('ms')
                break
        print(f'    {code}: IC={d["ic"]}, old_MS={old_ms}, new_MS(进面)={new_ms}')

# ══════════════════════════════════════════════════════════════
# Source 5: 韶关市直 笔试成绩 (2864709.xlsx) — WC+IC+MS
# ══════════════════════════════════════════════════════════════
def parse_shaoguan_zhi(data, by_code):
    fpath = find_file(3028707)
    if not fpath: print('韶关市直: XLSX not found'); return
    print('\n===== 韶关市直 笔试成绩 =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|岗位代码|准考证号|姓名|成绩|备注|排名
    pos_data = defaultdict(lambda: {'wc':0, 'ic':0, 'scores':[]})
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        code = str(row[1]).strip() if row[1] else ''
        if not code or code == 'None': continue
        d = pos_data[code]
        d['wc'] += 1
        score = safe_float(row[4])
        note = str(row[5]).strip() if row[5] else ''
        if score is not None: d['scores'].append(score)
        if '进入资格审核' in note or '进入资格复审' in note:
            d['ic'] += 1
    wb.close()

    updates = {}
    for code, d in pos_data.items():
        updates[code] = {
            'wc': d['wc'],
            'ic': d['ic'],
            'ms': min(d['scores']) if d['scores'] else None,
        }
    print(f'  Parsed: {len(updates)} codes, {sum(d["wc"] for d in pos_data.values())} candidates')
    apply_code_updates(data, by_code, updates, '韶关市直')

# ══════════════════════════════════════════════════════════════
# Source 6: 四会市 面试名单 (3259638.xlsx) — WC+MS+IC+PR
# ══════════════════════════════════════════════════════════════
def parse_sihui(data, by_code):
    fpath = find_file(3025405)
    if not fpath: print('四会: XLSX not found'); return
    print('\n===== 四会市 面试名单 =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|准考证号|姓名|性别|报考单位|报考职位|职位代码|招聘人数|笔试成绩|排名|是否通过资格审核|面试报到时间|备注
    pos_data = defaultdict(lambda: {'wc':0, 'scores':[], 'school':'', 'position':'', 'plan':0})
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        code = str(row[6]).strip() if row[6] else ''
        if not code or code == 'None': continue
        d = pos_data[code]
        d['school'] = str(row[4]).strip() if row[4] else ''
        d['position'] = str(row[5]).strip() if row[5] else ''
        d['plan'] = max(d['plan'], int(row[7]) if row[7] and str(row[7]).replace('.','').isdigit() else 0)
        d['wc'] += 1
        score = safe_float(row[8])
        if score is not None: d['scores'].append(score)
    wb.close()

    updates = {}
    for code, d in pos_data.items():
        updates[code] = {
            'wc': d['wc'], 'ic': d['wc'], 'pr': d['plan'],
            'ms': min(d['scores']) if d['scores'] else None,
        }
    print(f'  Parsed: {len(updates)} positions, {sum(d["wc"] for d in pos_data.values())} candidates')
    apply_code_updates(data, by_code, updates, '四会')

# ══════════════════════════════════════════════════════════════
# Source 7: 普宁市 面试名单 (1034039.xlsx) — IC
# ══════════════════════════════════════════════════════════════
def parse_puning(data, by_code):
    fpath = find_file(3020515)
    if not fpath: print('普宁: XLSX not found'); return
    print('\n===== 普宁市 面试名单 =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|准考证号|报考单位|报考职位|职位代码|面试日期|面试时间组别
    code_counts = defaultdict(int)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = str(row[4]).strip() if row[4] else ''
        if not code or code == 'None': continue
        code_counts[code] += 1
    wb.close()

    updates = {code: {'ic': count} for code, count in code_counts.items()}
    print(f'  Parsed: {len(updates)} codes, {sum(code_counts.values())} candidates')
    apply_code_updates(data, by_code, updates, '普宁')

# ══════════════════════════════════════════════════════════════
# Source 8: 浈江区 体检结果 (2864237.xlsx) — IC+PR
# ══════════════════════════════════════════════════════════════
def parse_zhenjiang(data, by_code):
    fpath = find_file(3023930)
    if not fpath: print('浈江: XLSX not found'); return
    print('\n===== 浈江区 体检结果 =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|职位代码|报考职位|招聘人数|准考证号|姓名|性别|排名|体检结果
    pos_data = defaultdict(lambda: {'total':0, 'passed':0, 'position':'', 'plan':0})
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = str(row[1]).strip() if row[1] else ''
        if not code or code == 'None': continue
        d = pos_data[code]
        d['position'] = str(row[2]).strip() if row[2] else ''
        d['plan'] = max(d['plan'], int(row[3]) if row[3] and str(row[3]).replace('.','').isdigit() else 0)
        d['total'] += 1
        result = str(row[8]).strip() if row[8] else ''
        if result == '合格': d['passed'] += 1
    wb.close()

    updates = {}
    for code, d in pos_data.items():
        updates[code] = {'ic': d['total'], 'pr': d['passed'], 'pr_override': True}
    print(f'  Parsed: {len(updates)} positions, {sum(d["total"] for d in pos_data.values())} total, {sum(d["passed"] for d in pos_data.values())} passed')
    apply_code_updates(data, by_code, updates, '浈江')

# ══════════════════════════════════════════════════════════════
# Source 9: 新丰县 体检结果 (2863719.xlsx + 2864785_1.xlsx) — IC+PR
# ══════════════════════════════════════════════════════════════
def parse_xinfeng(data, by_code):
    updates = {}
    for aid in [3018142, 3029753]:
        fpath = find_file(aid)
        if not fpath: continue
        print(f'\n===== 新丰县 体检结果 ({os.path.basename(fpath)}) =====')
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Cols: 序号|岗位代码|报考岗位|准考证号|性别|体检结果|备注
        pos_data = defaultdict(lambda: {'total':0, 'passed':0})
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            code = str(row[1]).strip() if row[1] else ''
            if not code or code == 'None': continue
            d = pos_data[code]
            d['total'] += 1
            result = str(row[5]).strip() if row[5] else ''
            if result == '合格': d['passed'] += 1
        wb.close()
        for code, d in pos_data.items():
            if code not in updates:
                updates[code] = {'total':0, 'passed':0}
            updates[code]['total'] += d['total']
            updates[code]['passed'] += d['passed']

    upds = {code: {'ic': d['total'], 'pr': d['passed'], 'pr_override': True}
            for code, d in updates.items()}
    print(f'  Combined: {len(upds)} positions, {sum(d["total"] for d in updates.values())} total')
    apply_code_updates(data, by_code, upds, '新丰')

# ══════════════════════════════════════════════════════════════
# Source 10: 连州市 体检合格 (2167533.xlsx) — IC+PR
# ══════════════════════════════════════════════════════════════
def parse_lianzhou(data, by_code):
    fpath = find_file(3021408)
    if not fpath: print('连州: XLSX not found'); return
    print('\n===== 连州市 体检合格 =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|职位代码|职位名称|准考证号|姓名|体检是否合格|是否进入考察
    pos_data = defaultdict(lambda: {'total':0, 'passed':0})
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        code = str(row[1]).strip() if row[1] else ''
        if not code or code == 'None': continue
        d = pos_data[code]
        d['total'] += 1
        passed = str(row[5]).strip() if row[5] else ''
        if passed in ('是', '合格'): d['passed'] += 1
    wb.close()

    updates = {}
    for code, d in pos_data.items():
        updates[code] = {'ic': d['total'], 'pr': d['passed'], 'pr_override': True}
    print(f'  Parsed: {len(updates)} positions, {sum(d["total"] for d in pos_data.values())} total')
    apply_code_updates(data, by_code, updates, '连州')

# ══════════════════════════════════════════════════════════════
# Source 11: 汕尾市直 面试资格审核 (1262235.xls) — IC
# ══════════════════════════════════════════════════════════════
def parse_shanwei_zhishu(data, by_code):
    fpath = find_file(3019065)
    if not fpath: print('汕尾市直: XLS not found'); return
    print('\n===== 汕尾市直 面试资格审核 =====')
    import xlrd
    wb = xlrd.open_workbook(fpath)
    ws = wb.sheet_by_index(0)
    code_counts = defaultdict(int)
    for i in range(2, ws.nrows):
        try:
            code_val = ws.cell_value(i, 2)
            if code_val:
                try: code = str(int(float(code_val)))
                except: code = str(code_val).strip()
                if code and code != 'None': code_counts[code] += 1
        except: pass
    wb.release_resources()

    updates = {code: {'ic': count} for code, count in code_counts.items()}
    print(f'  Parsed: {len(updates)} codes, {sum(code_counts.values())} candidates')
    apply_code_updates(data, by_code, updates, '汕尾市直')

# ══════════════════════════════════════════════════════════════
# Source 12: 汕尾城区 面试名单 (1262603.xlsx) — IC per school
# ══════════════════════════════════════════════════════════════
def parse_shanwei_chengqu(data, by_code):
    fpath = find_file(3022943)
    if not fpath: print('汕尾城区: XLSX not found'); return
    print('\n===== 汕尾城区 面试名单 (school-based matching) =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|准考证号|报考单位|面试组别
    school_counts = defaultdict(int)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        school = str(row[2]).strip() if row[2] else ''
        if not school or school == 'None': continue
        school = school.replace(' ', '').replace('\n', '')
        school_counts[school] += 1
    wb.close()

    print(f'  Schools: {len(school_counts)}, candidates: {sum(school_counts.values())}')

    # Match by school name
    matched = 0
    for i, r in enumerate(data):
        if r.get('c') != '汕尾': continue
        if r.get('d') not in ('城区', '汕尾城区'): continue
        sc = r.get('sc', '')
        for school, count in school_counts.items():
            if sc in school or school in sc:
                if r.get('ic', 0) == 0:
                    r['ic'] = count
                    stats['ic'] += 1
                    tag_current(r)
                    matched += 1
                break
    print(f'  School-matched IC: {matched} records')

# ══════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════
def main():
    print('=' * 70)
    print('V130 COMPREHENSIVE DATA FIX')
    print('=' * 70)

    data = load_data()
    print(f'Loaded: {len(data)} viewer records')
    by_code, by_key = build_indexes(data)
    print(f'Index: {len(by_code)} unique position codes')

    total_before = stats.copy()
    cur_before = sum(1 for r in data if r.get('_batch') == 'current')
    print(f'Currently tagged current: {cur_before}')

    # Phase 1: SCORES (WC, MS, TS)
    parse_xiashan(data, by_code)      # 霞山综合成绩
    parse_nanxiong(data, by_code)     # 南雄总成绩
    parse_shaoguan_zhi(data, by_code) # 韶关市直笔试

    # Phase 2: 经开区 special handling
    parse_jkq_written(data, by_code)  # 经开区笔试 → WC only (NO MS)
    parse_jkq_interview(data, by_code) # 经开区面试 → IC + CORRECT MS

    # Phase 3: INTERVIEW + PHYSICAL
    parse_sihui(data, by_code)        # 四会面试 (also WC+MS)
    parse_puning(data, by_code)       # 普宁面试 (IC)
    parse_zhenjiang(data, by_code)    # 浈江体检 (IC+PR)
    parse_xinfeng(data, by_code)      # 新丰体检 (IC+PR)
    parse_lianzhou(data, by_code)     # 连州体检 (IC+PR)
    parse_shanwei_zhishu(data, by_code)  # 汕尾市直面试 (IC)
    parse_shanwei_chengqu(data, by_code) # 汕尾城区面试 (IC)

    print(f'\n{"=" * 70}')
    print(f'FIX SUMMARY')
    print(f'{"=" * 70}')
    print(f'WC filled:    {stats["wc"]}')
    print(f'MS filled:    {stats["ms"]}')
    print(f'MS corrected: {stats["ms_corrected"]}')
    print(f'TS filled:    {stats["ts"]}')
    print(f'IC filled:    {stats["ic"]}')
    print(f'PR filled:    {stats["pr"]}')

    cur_after = sum(1 for r in data if r.get('_batch') == 'current')
    new_tagged = cur_after - cur_before
    print(f'\nNew records tagged: {new_tagged}')

    # Save
    save_data(data)

    # Audit 经开区
    jkq = [r for r in data if r.get('d') in ('经开区','湛江经开区') and r.get('c')=='湛江']
    jkq_wc = [r for r in jkq if r.get('wc',0) > 0]
    jkq_ic = [r for r in jkq if r.get('ic',0) > 0]
    jkq_ms = [r for r in jkq if r.get('ms') is not None and r.get('ms',0) > 0]
    print(f'\n经开区 post-fix: {len(jkq)} recs, WC>0={len(jkq_wc)}, IC>0={len(jkq_ic)}, MS>0={len(jkq_ms)}')
    if jkq_ms:
        print('MS values (sample):')
        for r in jkq_ms[:5]:
            print(f'  {r["sc"]} | {r["p"]} | WC={r["wc"]} IC={r["ic"]} MS={r["ms"]} (进面最低分)')

    # Audit 霞山
    xs = [r for r in data if r.get('d')=='霞山区' and r.get('c')=='湛江']
    xs_wc = sum(1 for r in xs if r.get('wc',0)>0)
    xs_ms = sum(1 for r in xs if r.get('ms') is not None and r.get('ms',0)>0)
    xs_ts = sum(1 for r in xs if r.get('ts') is not None and r.get('ts',0)>0)
    print(f'\n霞山区 post-fix: {len(xs)} recs, WC>0={xs_wc}, MS>0={xs_ms}, TS>0={xs_ts}')

if __name__ == '__main__':
    main()
