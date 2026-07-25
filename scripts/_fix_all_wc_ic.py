"""
COMPREHENSIVE WC==IC FIX:
1. 霞山 (2204552.xlsx 综合成绩): WC=total, IC=rows_with_interview_score
2. 南雄 (2863744.xlsx 总成绩): WC=total, IC=rows_with_interview_score
3. 乳源 (2864692_1.pdf 综合成绩): WC=total, IC=rows_with_interview_score
4. 四会 (面试名单): Source = interview only → WC should NOT be set
5. 陆河 (笔试名单): WC=IC IS correct (written exam list)
6. 浈江 (体检结果): IC=total, PR=passed — already correct structure

PERMANENT GUARD:
- Any record where WC>0 AND WC==IC: flag as suspicious
- Apply only to本次更新 records in this run
"""
import json, os, pdfplumber
from collections import defaultdict
import openpyxl, warnings
warnings.filterwarnings('ignore')

ROOT = r'D:\claude_code\gaokao\jiaozi\guangdong_scores'
VIEWER_PATH = os.path.join(ROOT, 'dist', 'viewer_data.json')
DL = os.path.join(ROOT, 'data', 'raw', 'tmp', '_gov_downloads_july22')

stats = {'wc_ok': 0, 'wc_fixed': 0, 'ic_fixed': 0, 'ms_fixed': 0, 'ts_fixed': 0, 'pr_fixed': 0}

def safe_float(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip())
    except: return None

def load_data():
    with open(VIEWER_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(data):
    with open(VIEWER_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

def find_dir(article_id):
    for d in os.listdir(DL):
        if str(article_id) in d:
            return os.path.join(DL, d)
    return None

def find_file(article_id, ext='.xlsx'):
    d = find_dir(article_id)
    if not d: return None
    for f in os.listdir(d):
        if f.endswith(ext) and not f.endswith('.xls' if ext == '.xlsx' else ''):
            return os.path.join(d, f)
    return None

def build_indexes(data):
    by_code = defaultdict(list)
    for i, r in enumerate(data):
        pc = str(r.get('position_code', '')).strip()
        if pc: by_code[pc].append(i)
    return by_code

def has_interview_score(row_data):
    """Check if a row has a valid interview score (not blank, not 缺考)"""
    s = str(row_data).strip() if row_data is not None else ''
    return s not in ('', '-', '缺考', 'None')

# ═══════════════════════════════════════════════════════════
# 1. 霞山区 — already fixed by _fix_xiashan_wc_ic.py
#    Re-running here for completeness
# ═══════════════════════════════════════════════════════════
def fix_xiashan(data, by_code):
    fpath = find_file(3025520)
    if not fpath:
        print('霞山: XLSX not found, skipping (already fixed)')
        return

    print('\n===== 霞山区 WC/IC fix =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Cols: 序号|岗位代码|招聘岗位|岗位名称|招聘人数|准考证号|笔试分数|面试分数|综合成绩|排名|是否入围体检
    per_code = defaultdict(lambda: {'wc': 0, 'ic': 0, 'written_i': [], 'composite_p': [], 'plan': 0})

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = str(row[1]).strip() if row[1] else ''
        if not code or code == 'None': continue
        d = per_code[code]
        d['wc'] += 1
        d['plan'] = max(d['plan'], int(row[4]) if row[4] and str(row[4]).replace('.', '').isdigit() else 0)
        written = safe_float(row[6])
        composite = safe_float(row[8])
        pass_check = str(row[10]).strip() if row[10] else ''

        if has_interview_score(row[7]):
            d['ic'] += 1
            if written is not None: d['written_i'].append(written)
        if composite is not None and '是' in pass_check:
            d['composite_p'].append(composite)
    wb.close()

    updates = {}
    for code, d in per_code.items():
        updates[code] = {
            'wc': d['wc'], 'ic': d['ic'], 'pr': d['plan'],
            'ms': round(min(d['written_i']), 2) if d['written_i'] else None,
            'ts': round(min(d['composite_p']), 2) if d['composite_p'] else None,
        }

    apply_fix(data, by_code, updates, '湛江', {'霞山区'}, '霞山')
    wc_eq = sum(1 for d in per_code.values() if d['wc'] == d['ic'])
    wc_gt = sum(1 for d in per_code.values() if d['wc'] > d['ic'])
    print(f'  XLSX: {len(per_code)} codes, WC==IC: {wc_eq}, WC>IC: {wc_gt}')

# ═══════════════════════════════════════════════════════════
# 2. 南雄市 — same bug as 霞山
# ═══════════════════════════════════════════════════════════
def fix_nanxiong(data, by_code):
    fpath = find_file(3018354)
    if not fpath:
        print('南雄: XLSX not found')
        return

    print('\n===== 南雄市 WC/IC fix =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb['Sheet1']
    # Cols: 序号|准考证号|岗位代码|岗位名称|招聘单位|招聘人数|笔试成绩|面试成绩|总成绩|排名|是否进入体检|备注
    per_code = defaultdict(lambda: {'wc': 0, 'ic': 0, 'written_i': [], 'composite_p': [], 'plan': 0})

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        code = str(row[2]).strip() if row[2] else ''
        if not code or code == 'None': continue
        d = per_code[code]
        d['wc'] += 1
        d['plan'] = max(d['plan'], int(row[5]) if row[5] and str(row[5]).replace('.', '').isdigit() else 0)
        written = safe_float(row[6])
        total = safe_float(row[8])
        pass_check = str(row[10]).strip() if row[10] else ''

        if has_interview_score(row[7]):  # 面试成绩 column
            d['ic'] += 1
            if written is not None: d['written_i'].append(written)
        if total is not None and '是' in pass_check:
            d['composite_p'].append(total)
    wb.close()

    updates = {}
    for code, d in per_code.items():
        updates[code] = {
            'wc': d['wc'], 'ic': d['ic'], 'pr': d['plan'],
            'ms': round(min(d['written_i']), 2) if d['written_i'] else None,
            'ts': round(min(d['composite_p']), 2) if d['composite_p'] else None,
        }

    apply_fix(data, by_code, updates, '韶关', {'南雄市'}, '南雄')
    wc_eq = sum(1 for d in per_code.values() if d['wc'] == d['ic'])
    wc_gt = sum(1 for d in per_code.values() if d['wc'] > d['ic'])
    print(f'  XLSX: {len(per_code)} codes, WC==IC: {wc_eq}, WC>IC: {wc_gt}')

# ═══════════════════════════════════════════════════════════
# 3. 乳源 — PDF 综合成绩, same bug
# ═══════════════════════════════════════════════════════════
def fix_ruyuan(data, by_code):
    # Article 3028394, file 2864692_1.pdf (the one with笔试+面试 detail columns)
    dirpath = find_dir(3028394)
    fpath = None
    if dirpath:
        for f in os.listdir(dirpath):
            if f == '2864692_1.pdf':
                fpath = os.path.join(dirpath, f)
                break
    if not fpath:
        print('乳源: 2864692_1.pdf not found')
        return

    print(f'\n===== 乳源县 WC/IC fix =====')
    per_code = defaultdict(lambda: {'wc': 0, 'ic': 0, 'written_i': [], 'composite_p': [], 'plan': 0})

    with pdfplumber.open(fpath) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or not table[0]: continue
                for row in table[1:]:
                    if not row or not row[3]: continue
                    code = str(row[3]).strip().replace('\n', '') if row[3] else ''
                    if not code or code == 'None': continue
                    d = per_code[code]
                    d['wc'] += 1
                    d['plan'] = max(d['plan'], int(row[4]) if row[4] and str(row[4]).strip().replace('.', '').isdigit() else 0)
                    written = safe_float(row[6])
                    composite = safe_float(row[8])
                    pass_check = str(row[10]).strip() if row[10] else ''

                    # Check if has interview score (col 7) or composite score
                    has_intv = (row[7] is not None and str(row[7]).strip() not in ('', '-', '缺考'))
                    if has_intv:
                        d['ic'] += 1
                        if written is not None: d['written_i'].append(written)
                    if composite is not None and '是' in pass_check:
                        d['composite_p'].append(composite)

    updates = {}
    for code, d in per_code.items():
        updates[code] = {
            'wc': d['wc'], 'ic': d['ic'], 'pr': d['plan'],
            'ms': round(min(d['written_i']), 2) if d['written_i'] else None,
            'ts': round(min(d['composite_p']), 2) if d['composite_p'] else None,
        }

    apply_fix(data, by_code, updates, '韶关', {'乳源瑶族自治县', '乳源县'}, '乳源')
    wc_eq = sum(1 for d in per_code.values() if d['wc'] == d['ic'])
    wc_gt = sum(1 for d in per_code.values() if d['wc'] > d['ic'])
    print(f'  PDF: {len(per_code)} codes, WC==IC: {wc_eq}, WC>IC: {wc_gt}')

# ═══════════════════════════════════════════════════════════
# Helper: apply corrected data
# ═══════════════════════════════════════════════════════════
def apply_fix(data, by_code, updates, city, districts, label):
    applied = 0
    for code, upd in updates.items():
        code = str(code).strip()
        if code not in by_code: continue
        for idx in by_code[code]:
            r = data[idx]
            if r.get('c') != city: continue
            if r.get('d') not in districts: continue

            changed = False

            # WC: set if > 0 and different
            if upd['wc'] > 0 and r.get('wc', 0) != upd['wc']:
                r['wc'] = upd['wc']; stats['wc_fixed'] += 1; changed = True

            # IC: set if > 0 and different
            if upd['ic'] > 0 and r.get('ic', 0) != upd['ic']:
                r['ic'] = upd['ic']; stats['ic_fixed'] += 1; changed = True

            # MS: overwrite if different (进面最低分)
            if upd['ms'] is not None and r.get('ms') != upd['ms']:
                r['ms'] = upd['ms']; stats['ms_fixed'] += 1; changed = True

            # TS: set if new or different
            if upd['ts'] is not None:
                if r.get('ts') is None or r.get('ts', 0) == 0:
                    r['ts'] = upd['ts']; stats['ts_fixed'] += 1; changed = True
                elif r.get('ts') != upd['ts']:
                    r['ts'] = upd['ts']; stats['ts_fixed'] += 1; changed = True

            # PR: 体检合格 > 计划数
            if upd['pr'] > 0:
                if r.get('pr', 0) == 0:
                    r['pr'] = upd['pr']; stats['pr_fixed'] += 1; changed = True
                elif upd['pr'] > r.get('pr', 0):
                    r['pr'] = upd['pr']; stats['pr_fixed'] += 1; changed = True

            if changed:
                r['_batch'] = 'current'
                r['_batch_label'] = '本次更新'
                if r.get('wc', 0) > 0 and r.get('pr', 0) > 0:
                    r['wr'] = round(r['wc'] / r['pr'], 2)
                if r.get('ic', 0) > 0 and r.get('pr', 0) > 0:
                    r['ir'] = round(r['ic'] / r['pr'], 2)
                applied += 1

    print(f'  [{label}] {applied} record updates')

def main():
    data = load_data()
    print(f'Loaded: {len(data)} records')
    by_code = build_indexes(data)

    fix_xiashan(data, by_code)
    fix_nanxiong(data, by_code)
    fix_ruyuan(data, by_code)

    # ═══════════════════════════════════════════════════════════
    # GLOBAL AUDIT
    # ═══════════════════════════════════════════════════════════
    print('\n' + '=' * 60)
    print('SUMMARY')
    print('=' * 60)
    print(f'WC fixed: {stats["wc_fixed"]}')
    print(f'IC fixed: {stats["ic_fixed"]}')
    print(f'MS fixed: {stats["ms_fixed"]}')
    print(f'TS fixed: {stats["ts_fixed"]}')
    print(f'PR fixed: {stats["pr_fixed"]}')

    # Audit本次更新 WC==IC
    current = [r for r in data if r.get('_batch') == 'current']
    cur_wc_ic = [(r, f"{r.get('c','')}/{r.get('d','')}") for r in current if r.get('wc',0) > 0 and r.get('wc',0) == r.get('ic',0)]

    by_dist = defaultdict(list)
    for r, dist in cur_wc_ic:
        by_dist[dist].append(r)

    print(f'\n本次更新 WC==IC 剩余: {len(cur_wc_ic)}条')
    for k in sorted(by_dist, key=lambda k: -len(by_dist[k])):
        print(f'  {k}: {len(by_dist[k])}条')

    # Show problematic ones
    if cur_wc_ic:
        print('\n需要关注的剩余 WC==IC:')
        sus = [(r, d) for r, d in cur_wc_ic if d not in ('汕尾/陆河县',)]  # 陆河笔试名单 OK
        sus2 = [(r, d) for r, d in sus if d not in ('湛江/霞山区', '韶关/南雄市', '韶关/乳源瑶族自治县', '韶关/乳源县')]
        for r, d in sus2[:15]:
            print(f'  {d}: WC={r["wc"]} IC={r["ic"]} MS={r.get("ms")} code={r.get("position_code","")} | {r.get("sc","")} | {r.get("p","")} | yr={r.get("yr","")}')

    # Per-district stats for fixed districts
    for city, dist in [('湛江','霞山区'), ('韶关','南雄市'), ('韶关','乳源瑶族自治县')]:
        recs = [r for r in data if r.get('c')==city and r.get('d')==dist]
        wc_gt_0 = sum(1 for r in recs if r.get('wc',0)>0)
        wc_eq_ic = sum(1 for r in recs if r.get('wc',0)>0 and r.get('wc',0)==r.get('ic',0))
        wc_gt_ic = sum(1 for r in recs if r.get('wc',0)>0 and r.get('wc',0)>r.get('ic',0))
        print(f'\n{city}/{dist}: {len(recs)}条, WC>0={wc_gt_0}, WC==IC={wc_eq_ic}, WC>IC={wc_gt_ic}')

    # ═══════════════════════════════════════════════════════════
    # PERMANENT GUARD: Add WC==IC validation layer
    # Write guard metadata to data itself
    # ═══════════════════════════════════════════════════════════
    for r in data:
        wc, ic = r.get('wc', 0), r.get('ic', 0)
        if wc > 0 and wc == ic:
            # Mark source so viewer/future scripts can flag
            r['_wc_ic_same_source'] = True
        else:
            r['_wc_ic_same_source'] = False

    save_data(data)
    print(f'\nSaved: {len(data)} records')
    print('WC==IC guard metadata written (_wc_ic_same_source)')

if __name__ == '__main__':
    main()
